from flask import Flask, request
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)
EXCEL_FILE = "donnees.xlsx"

def save_to_excel(new_data: dict):
    # Create workbook if not exists
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        # write header row
        headers = list(new_data.keys())
        ws.append(headers)
        wb.save(EXCEL_FILE)

    # Load and ensure columns cover all keys
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    existing_headers = [cell.value for cell in ws[1]]
    # add any missing headers (new keys)
    for k in new_data.keys():
        if k not in existing_headers:
            existing_headers.append(k)
            ws.cell(row=1, column=len(existing_headers)).value = k

    # build row in header order
    row = [new_data.get(h, "") for h in existing_headers]
    ws.append(row)
    wb.save(EXCEL_FILE)

@app.route('/receive', methods=['POST'])
def receive_data():
    data = request.json
    if not data or not isinstance(data, dict):
        return {"status": "error", "message": "No JSON received"}, 400
    save_to_excel(data)
    return {"status": "success", "message": "Data saved to Excel"}

if __name__ == '__main__':
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
